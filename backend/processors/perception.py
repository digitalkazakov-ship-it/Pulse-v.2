import openpyxl

BRANDS = ['Narzan', 'Borjomi', 'Senezhskaya', 'Svyatoy']
BRAND_DISPLAY = {
    'Narzan': 'Нарзан', 'Borjomi': 'Боржоми',
    'Senezhskaya': 'Сенежская', 'Svyatoy': 'Святой источник',
}
RU_MONTHS = {1:'Янв',2:'Фев',3:'Мар',4:'Апр',5:'Май',6:'Июн',7:'Июл',8:'Авг',9:'Сен',10:'Окт',11:'Ноя',12:'Дек'}

POS_COLS = {'Narzan': 3, 'Svyatoy': 5, 'Senezhskaya': 6, 'Borjomi': 8}
NEG_COLS = {'Narzan': 15, 'Svyatoy': 17, 'Senezhskaya': 18, 'Borjomi': 20}
ABS_COLS = {'Narzan': 2, 'Borjomi': 3, 'Svyatoy': 4, 'Senezhskaya': 6}
MED_COLS = {'Narzan': 22, 'Svyatoy': 24, 'Senezhskaya': 25, 'Borjomi': 27}
ALL_ROWS  = list(range(2, 22))
CHAR_ROWS = [2, 4, 5, 8, 10, 18, 20]

SL_CATEGORIES = [
    {'key': 'assortment', 'label': 'Ассортимент', 'params': [
        {'key': 'p0', 'label': 'Уникальные и эксклюзивные продукты', 'row': 4},
        {'key': 'p1', 'label': 'Широкий ассортимент', 'row': 5},
        {'key': 'p2', 'label': 'Высокое качество товаров', 'row': 6},
        {'key': 'p3', 'label': 'Высокое качество упаковки', 'row': 7},
    ]},
    {'key': 'price', 'label': 'Цена', 'params': [
        {'key': 'p0', 'label': 'Конкурентоспособные и выгодные цены', 'row': 8},
        {'key': 'p1', 'label': 'Наличие скидок и промо', 'row': 9},
    ]},
    {'key': 'product', 'label': 'Продукт', 'params': [
        {'key': 'p0', 'label': 'Приятный вкус', 'row': 13},
        {'key': 'p1', 'label': 'Польза от употребления', 'row': 14},
        {'key': 'p2', 'label': 'Можно часто пить', 'row': 15},
    ]},
]


def _int_val(v): return int(v) if isinstance(v, (int, float)) else 0


def process(file_path: str, generated: str = '') -> dict:
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    rows_sl    = list(wb['Шаг 4.1 - SL'].iter_rows(values_only=True))
    rows_image = list(wb['Шаг 4.4 - Имидж'].iter_rows(values_only=True))
    wb.close()

    sl: dict = {}
    pos_total = {b: 0 for b in BRANDS}
    neg_total = {b: 0 for b in BRANDS}

    for cat in SL_CATEGORIES:
        chart_data: dict = {}
        cat_pos = {b: 0 for b in BRANDS}
        cat_neg = {b: 0 for b in BRANDS}
        for p in cat['params']:
            row = rows_sl[p['row']]
            pts = []
            for b in BRANDS:
                pv = _int_val(row[POS_COLS[b]])
                nv = _int_val(row[NEG_COLS[b]])
                pts.append({'brand': b, 'brandName': BRAND_DISPLAY[b], 'pos': pv, 'neg': nv})
                cat_pos[b] += pv
                cat_neg[b] += nv
            chart_data[p['key']] = pts
        chart_data['total'] = [
            {'brand': b, 'brandName': BRAND_DISPLAY[b], 'pos': cat_pos[b], 'neg': cat_neg[b]}
            for b in BRANDS
        ]
        for b in BRANDS:
            pos_total[b] += cat_pos[b]
            neg_total[b] += cat_neg[b]
        sl[cat['key']] = {
            'label': cat['label'],
            'params': [{'key': 'total', 'label': 'Все параметры'}] + [{'key': p['key'], 'label': p['label']} for p in cat['params']],
            'chartData': chart_data,
        }

    sentiment = {
        b: round((pos_total[b] - neg_total[b]) / (pos_total[b] + neg_total[b]) * 100, 1)
        if (pos_total[b] + neg_total[b]) > 0 else 0.0
        for b in BRANDS
    }

    def _abs_point(r):
        row = rows_image[r]
        return {'label': str(row[1]), **{b: _int_val(row[ABS_COLS[b]]) for b in BRANDS}}

    def _med_point(r):
        row = rows_image[r]
        pt = {'label': str(row[1])}
        for b in BRANDS:
            v = row[MED_COLS[b]]
            pt[b] = round(float(v) * 100, 1) if isinstance(v, (int, float)) else 0.0
        return pt

    # Determine period label
    if generated:
        y, m = int(generated[:4]), int(generated[5:7])
        lbl = RU_MONTHS[m] + (f" '{str(y)[2:]}" if y != 2026 else '')
        sentiment_entry = {'period': generated, 'month': lbl, **sentiment}
    else:
        sentiment_entry = {'period': '', 'month': '', **sentiment}

    return {
        'generated':  generated,
        'brands':     BRANDS,
        'brandNames': BRAND_DISPLAY,
        'sl':         sl,
        'sentiment':  [sentiment_entry],
        'imageChars': [_abs_point(r) for r in CHAR_ROWS],
        'matrix': {
            'absolute': [_abs_point(r) for r in ALL_ROWS],
            'median':   [_med_point(r) for r in ALL_ROWS],
        },
    }
