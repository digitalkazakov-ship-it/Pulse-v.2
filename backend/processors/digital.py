"""
Multi-file processor for digital_audit.json.
Expects a ZIP containing exactly two .xlsx files:
  - current year  (e.g. 'report act.xlsx')
  - previous year (e.g. 'report act 2.xlsx')
Files are sorted alphabetically; the first is treated as current year,
the second as previous year.
"""
import os
import glob
import datetime
from collections import defaultdict
import openpyxl

BRANDS = ['BrandX', 'CompA', 'CompB', 'CompC']
BRAND_MAP = {
    'theact.ru':    'BrandX',
    'librederm.ru': 'CompA',
    'mixit.ru':     'CompB',
    'vois.ru':      'CompC',
}
BRAND_DISPLAY = {
    'BrandX': 'The Act',
    'CompA':  'Librederm',
    'CompB':  'Mixit',
    'CompC':  'Vois',
}

RU_MONTHS = {
    '01': 'Янв', '02': 'Фев', '03': 'Мар', '04': 'Апр',
    '05': 'Май', '06': 'Июн', '07': 'Июл', '08': 'Авг',
    '09': 'Сен', '10': 'Окт', '11': 'Ноя', '12': 'Дек',
}

CURR_YEAR = str(datetime.date.today().year)


def _month_label(ym: str) -> str:
    year, month = ym.split('-')
    label = RU_MONTHS[month]
    if year != CURR_YEAR:
        label += f" '{year[2:]}"
    return label


def _find_files(dir_path: str) -> tuple[str, str]:
    xlsx = sorted(glob.glob(os.path.join(dir_path, '**/*.xlsx'), recursive=True))
    if len(xlsx) < 2:
        raise FileNotFoundError(
            f"Expected 2 .xlsx files in ZIP, found {len(xlsx)}"
        )
    return xlsx[0], xlsx[1]


def _read_daily_sheet(wb, sheet_name: str, agg: str = 'sum') -> dict:
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    header = rows[0]
    result: dict = {}
    for row in rows[1:]:
        if not row[0]:
            continue
        site = str(row[0]).strip()
        if site not in BRAND_MAP:
            continue
        brand = BRAND_MAP[site]
        buckets: dict = defaultdict(list)
        for col_i, val in enumerate(row[1:], start=1):
            if col_i >= len(header):
                break
            date_val = header[col_i]
            if date_val is None:
                continue
            ym = str(date_val)[:7]
            if isinstance(val, (int, float)) and val > 0:
                buckets[ym].append(val)
        result[brand] = {}
        for ym, vals in buckets.items():
            result[brand][ym] = sum(vals) if agg == 'sum' else sum(vals) / len(vals)
    return result


def _read_unique_visits(wb) -> dict:
    ws = wb['Unique_Visits']
    rows = list(ws.iter_rows(values_only=True))
    header = rows[0]
    months = [str(h)[:7] for h in header[2:] if h is not None]
    totals: dict = defaultdict(lambda: defaultdict(float))
    for row in rows[1:]:
        if not row[0]:
            continue
        site = str(row[0]).strip()
        if site not in BRAND_MAP:
            continue
        brand = BRAND_MAP[site]
        for i, ym in enumerate(months):
            val = row[2 + i]
            if isinstance(val, (int, float)) and val > 0:
                totals[brand][ym] += val
    return {b: dict(m) for b, m in totals.items()}


def _read_traffic_sources(wb, sheet_name: str, device: str) -> dict:
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    header = rows[0]
    date_start = 3 if device == 'desktop' else 2
    date_cols = [h for h in header[date_start:] if h is not None]
    if not date_cols:
        return {}
    latest_idx = date_start + len(date_cols) - 1

    PAID_SOURCES = {'Display Ads', 'Paid Search'}
    SOURCE_LABEL = {
        'Direct': 'Direct', 'Organic Search': 'Search', 'Paid Search': 'Search',
        'Display Ads': 'Display Ads', 'Social Media': 'Social',
        'Referral': 'Referral', 'Mail': 'Mail',
    }

    brand_buckets: dict = defaultdict(lambda: defaultdict(lambda: {'organic': 0.0, 'paid': 0.0}))
    for row in rows[1:]:
        if not row[0]:
            continue
        site = str(row[0]).strip()
        if site not in BRAND_MAP:
            continue
        brand = BRAND_MAP[site]
        source_raw = str(row[1]).strip() if row[1] else ''
        label = SOURCE_LABEL.get(source_raw, source_raw)
        val = row[latest_idx] if latest_idx < len(row) else None
        if not isinstance(val, (int, float)) or val <= 0:
            continue
        if device == 'desktop':
            visit_type = str(row[2]).strip() if row[2] else ''
            if 'paid' in visit_type.lower() or source_raw in PAID_SOURCES:
                brand_buckets[brand][label]['paid'] += val
            else:
                brand_buckets[brand][label]['organic'] += val
        else:
            if source_raw in PAID_SOURCES:
                brand_buckets[brand][label]['paid'] += val
            else:
                brand_buckets[brand][label]['organic'] += val

    result: dict = {}
    for brand, buckets in brand_buckets.items():
        rows_out = []
        for source, vals in buckets.items():
            total = vals['organic'] + vals['paid']
            if total > 0:
                rows_out.append({'source': source, 'organic': round(vals['organic']),
                                 'paid': round(vals['paid']), 'total': round(total)})
        rows_out.sort(key=lambda x: -x['total'])
        result[brand] = rows_out
    return result


def _merge(d1: dict, d2: dict) -> dict:
    out: dict = {}
    for brand in set(d1) | set(d2):
        out[brand] = {}
        if brand in d2:
            out[brand].update(d2[brand])
        if brand in d1:
            out[brand].update(d1[brand])
    return out


def _trend(data: dict, months: list, scale: float = 1.0, decimals: int = 1) -> list:
    return [
        {'month': _month_label(ym),
         **{b: round(data.get(b, {}).get(ym, 0) * scale, decimals) for b in BRANDS}}
        for ym in months
    ]


def _detect_latest_month(data: dict) -> str:
    months: set = set()
    for brand_data in data.values():
        months.update(brand_data.keys())
    return max(months) if months else ''


def process(dir_path: str) -> dict:
    curr_file, prev_file = _find_files(dir_path)

    wbc = openpyxl.load_workbook(curr_file, read_only=True, data_only=True)
    wbp = openpyxl.load_workbook(prev_file, read_only=True, data_only=True)

    visits = _merge(_read_daily_sheet(wbc, 'Visits', 'sum'),
                    _read_daily_sheet(wbp, 'Visits', 'sum'))
    unique = _merge(_read_unique_visits(wbc), _read_unique_visits(wbp))
    pages  = _merge(_read_daily_sheet(wbc, 'Pages_Per_Visit', 'mean'),
                    _read_daily_sheet(wbp, 'Pages_Per_Visit', 'mean'))
    dur    = _merge(_read_daily_sheet(wbc, 'Avg_Visit_Duration', 'mean'),
                    _read_daily_sheet(wbp, 'Avg_Visit_Duration', 'mean'))
    bounce = _merge(_read_daily_sheet(wbc, 'Bounce_Rate', 'mean'),
                    _read_daily_sheet(wbp, 'Bounce_Rate', 'mean'))

    td = _read_traffic_sources(wbc, 'Desktop_Overview_Share', 'desktop')
    tm = _read_traffic_sources(wbc, 'Mobile_Overview_Share',  'mobile')

    wbc.close()
    wbp.close()

    curr = _detect_latest_month(visits)
    if not curr:
        return {'generated': '', 'brands': BRANDS, 'brandNames': BRAND_DISPLAY,
                'metrics': [], 'visitsTrend': [], 'trafficSourcesDesktop': {},
                'trafficSourcesMobile': {}}

    curr_year, curr_mo = curr.split('-')
    prev_mo = f"{curr_year}-{int(curr_mo)-1:02d}" if int(curr_mo) > 1 \
              else f"{int(curr_year)-1}-12"
    prev_year_same_mo = f"{int(curr_year)-1}-{curr_mo}"

    # 6-month and 12-month trailing windows
    def trailing(n: int) -> list:
        y, m = int(curr_year), int(curr_mo)
        out = []
        for _ in range(n):
            out.insert(0, f"{y}-{m:02d}")
            m -= 1
            if m == 0:
                m = 12
                y -= 1
        return out

    t6  = trailing(6)
    t12 = trailing(12)

    def bvals(data, ym, scale=1.0, dec=1):
        return {b: round(data.get(b, {}).get(ym, 0) * scale, dec) for b in BRANDS}

    def mom_yoy(data, scale=1.0):
        c = data.get('BrandX', {}).get(curr, 0) * scale
        p = data.get('BrandX', {}).get(prev_mo, 0) * scale
        y = data.get('BrandX', {}).get(prev_year_same_mo, 0) * scale
        mom = round((c - p) / p * 100, 1) if p else 0
        yoy = round((c - y) / y * 100, 1) if y else 0
        return mom, yoy

    v_mom, v_yoy = mom_yoy(visits, 1/1000)
    u_mom, u_yoy = mom_yoy(unique, 1/1000)
    p_mom, p_yoy = mom_yoy(pages)
    d_mom, d_yoy = mom_yoy(dur, 1/60)
    b_mom, b_yoy = mom_yoy(bounce, 100)

    metrics = [
        {'key': 'visits',   'metric': 'Monthly Visits (K)',    **bvals(visits,  curr, 1/1000),    'mom': v_mom, 'yoy': v_yoy, 'trend': _trend(visits,  t6, 1/1000)},
        {'key': 'uniques',  'metric': 'Unique Visitors (K)',   **bvals(unique,  curr, 1/1000),    'mom': u_mom, 'yoy': u_yoy, 'trend': _trend(unique,  t6, 1/1000)},
        {'key': 'pages',    'metric': 'Pages/Visit',           **bvals(pages,   curr, dec=2),     'mom': p_mom, 'yoy': p_yoy, 'trend': _trend(pages,   t6, decimals=2)},
        {'key': 'duration', 'metric': 'Avg Duration (min)',    **bvals(dur,     curr, 1/60),      'mom': d_mom, 'yoy': d_yoy, 'trend': _trend(dur,     t6, 1/60)},
        {'key': 'bounce',   'metric': 'Bounce Rate (%)',       **bvals(bounce,  curr, 100),       'mom': b_mom, 'yoy': b_yoy, 'trend': _trend(bounce,  t6, 100)},
    ]

    return {
        'generated':             curr,
        'brands':                BRANDS,
        'brandNames':            BRAND_DISPLAY,
        'metrics':               metrics,
        'visitsTrend':           _trend(visits, t12, 1/1000),
        'trafficSourcesDesktop': td,
        'trafficSourcesMobile':  tm,
    }
