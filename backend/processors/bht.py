import re
import openpyxl
from datetime import datetime as _dt

# ── Mineral-water constants ────────────────────────────────────────────────────

BRANDS = ['Narzan', 'Borjomi', 'Senezhskaya', 'Svyatoy']
BRAND_DISPLAY = {
    'Narzan': 'Нарзан', 'Borjomi': 'Боржоми',
    'Senezhskaya': 'Сенежская', 'Svyatoy': 'Святой источник',
}
QUARTERS  = ['Q1 2026', 'Q4 2025', 'Q3 2025', 'Q2 2025']
CURRENT_Q = 'Q1 2026'
PRIOR_Q   = 'Q4 2025'

TOTAL_ROW_S0 = 10
METRIC_ROWS = {
    'topOfMind':     {'Borjomi': 12, 'Narzan': 13, 'Svyatoy': 15, 'Senezhskaya': 16},
    'spontaneous':   {'Borjomi': 18, 'Narzan': 19, 'Svyatoy': 21, 'Senezhskaya': 22},
    'aided':         {'Borjomi': 24, 'Narzan': 25, 'Svyatoy': 27, 'Senezhskaya': 28},
    'consumption':   {'Borjomi': 30, 'Narzan': 31, 'Svyatoy': 33, 'Senezhskaya': 34},
    'consideration': {'Borjomi': 42, 'Narzan': 43, 'Svyatoy': 45, 'Senezhskaya': 46},
}
METRIC_KEYS   = ['topOfMind', 'spontaneous', 'aided', 'consideration', 'consumption']
METRIC_LABELS = {
    'topOfMind': 'Первое упоминание', 'spontaneous': 'Спонтанное',
    'aided': 'Подсказанное', 'consideration': 'Рассмотрение', 'consumption': 'Потребление',
}
CHRONO = [('Q2 2025', 5), ('Q3 2025', 4), ('Q4 2025', 3), ('Q1 2026', 2)]

AGE_TOTAL_ROW    = 9
CONSUMPTION_ROWS = {'Borjomi': 29, 'Narzan': 30, 'Svyatoy': 32, 'Senezhskaya': 33}
AGE_GROUPS       = ['Total', '12-17', '18-24', '25-34', '35-44', '45-54', '55-64']
AGE_COL_START    = {'Total': 2, '12-17': 10, '18-24': 18, '25-34': 26, '35-44': 34, '45-54': 42, '55-64': 50}

FREQ_TOTAL_ROW = 9
FREQ_ROWS = {'Borjomi': 11, 'Narzan': 12, 'Svyatoy': 14, 'Senezhskaya': 15}
FREQ_CATS  = ['daily', 'week23', 'week1', 'month23', 'less']
FREQ_LABELS = {
    'daily': 'Каждый день', 'week23': '2-3 раза в нед.',
    'week1': '1 раз в нед.', 'month23': '2-3 раза в мес.', 'less': 'Еще реже',
}
FREQ_COL_START = {'total': 2, 'daily': 6, 'week23': 10, 'week1': 14, 'month23': 18, 'less': 22}

# ── Monthly Brand Pulse format (universal auto-detector) ──────────────────────

_METRIC_LABEL_TO_KEY = {
    'Первое упоминание':           'topOfMind',
    'Спонтанное знание':           'spontaneous',
    'Подсказанное знание':         'aided',
    'Рассмотрение к покупке':      'consideration',
    'Потребление':                 'consumption',
    'Наиболее частое потребление': 'mostFrequent',
}
_METRIC_KEY_ORDER = ['topOfMind', 'spontaneous', 'aided', 'consideration', 'consumption', 'mostFrequent']

_RU_MONTHS = ['янв', 'фев', 'мар', 'апр', 'май', 'июн', 'июл', 'авг', 'сен', 'окт', 'ноя', 'дек']

# ── Helpers ────────────────────────────────────────────────────────────────────

def vf(v):
    return round(float(v), 1) if isinstance(v, (int, float)) else 0.0


def _cv(v):
    if v == '*' or v is None:
        return None
    try:
        return round(float(v), 2)
    except (TypeError, ValueError):
        return None


def _brand_key(display_name: str) -> str:
    """Sanitize brand display name into a compact key (removes spaces and punctuation)."""
    return re.sub(r'[^\w]', '', display_name, flags=re.UNICODE)


# Col 0 section-header keywords → segment group
_PEN_SEG_KEYWORDS = {
    'пол':      'gender',
    'возраст':  'age',
    'размер':   'citySize',
    'федерал':  'federalDistrict',
}


def _is_monthly_format(wb) -> bool:
    """Monthly Brand Pulse files have datetime values in row 12 (1-indexed) of sheet 0."""
    row = list(wb.worksheets[0].iter_rows(min_row=12, max_row=12, values_only=True))[0]
    return any(isinstance(v, _dt) for v in row)


def _process_monthly_auto(wb) -> dict:
    rows0 = list(wb.worksheets[0].iter_rows(values_only=True))

    # Penetration sheet: the first sheet whose name contains 'пенетрация'
    pen_ws = next((ws for ws in wb.worksheets if 'пенетрация' in ws.title.lower()), None)

    # Month → column index from row 11 (0-indexed) of the voronki sheet.
    # The sheet repeats months for each age-group interval — stop at first repeated month
    # so we only use the first (total) interval.
    month_col: dict = {}
    for ci, v in enumerate(rows0[11]):
        if isinstance(v, _dt):
            key = f'{v.year}-{v.month:02d}'
            if key in month_col:
                break  # second interval starts here
            month_col[key] = ci
    months_sorted = sorted(month_col.keys())

    def _qlabel(ym: str) -> str:
        y, m = int(ym[:4]), int(ym[5:])
        return f'Q{(m - 1) // 3 + 1} {y}'

    def _mlabel(ym: str) -> str:
        y, m = int(ym[:4]), int(ym[5:])
        return f'{_RU_MONTHS[m - 1]} {y}'

    # Detect brands: rows where col[0]='Первое упоминание' and col[1] is the brand display name.
    # Within each brand block (up to 12 rows) scan for all known metric labels.
    brands_order: list = []
    brand_display: dict = {}
    brand_metric_rows: dict = {}  # {bk: {metric_key: row_index}}

    for ri in range(12, min(len(rows0), 300)):
        row = rows0[ri]
        if not row or len(row) < 2:
            continue
        col0, col1 = row[0], row[1]
        if (isinstance(col0, str) and col0.strip() == 'Первое упоминание' and
                isinstance(col1, str) and col1.strip()):
            bdisp = col1.strip()
            bk = _brand_key(bdisp)
            if bk not in brand_display:
                brands_order.append(bk)
                brand_display[bk] = bdisp
                brand_metric_rows[bk] = {}
            for sub_ri in range(ri, ri + 12):
                if sub_ri >= len(rows0):
                    break
                sub_row = rows0[sub_ri]
                if not sub_row:
                    continue
                lbl = sub_row[0]
                if isinstance(lbl, str):
                    mk = _METRIC_LABEL_TO_KEY.get(lbl.strip())
                    if mk:
                        brand_metric_rows[bk][mk] = sub_ri

    all_metric_set: set = set()
    for bk in brands_order:
        all_metric_set |= set(brand_metric_rows[bk].keys())
    metric_keys = [k for k in _METRIC_KEY_ORDER if k in all_metric_set]
    metric_labels_out = {k: lbl for lbl, k in _METRIC_LABEL_TO_KEY.items()}

    # Build funnel metrics time series
    metrics: dict = {}
    for mk in metric_keys:
        series = []
        for ym in months_sorted:
            ci = month_col[ym]
            entry: dict = {'month': ym, 'monthLabel': _mlabel(ym), 'quarter': _qlabel(ym)}
            for bk in brands_order:
                ri = brand_metric_rows[bk].get(mk)
                if ri is not None and ri < len(rows0) and len(rows0[ri]) > ci:
                    entry[bk] = _cv(rows0[ri][ci])
                else:
                    entry[bk] = None
            series.append(entry)
        metrics[mk] = series

    # Penetration segments
    pen_segments: dict = {'total': [], 'gender': [], 'age': [], 'citySize': [], 'federalDistrict': []}
    years: list = []
    brand_col_starts: dict = {}

    if pen_ws and brands_order:
        pen_rows = list(pen_ws.iter_rows(values_only=True))

        # Find year row: first row after row 30 with ≥3 year integers (2015-2030)
        year_row_i = -1
        for ri in range(30, min(len(pen_rows), 80)):
            row = pen_rows[ri]
            if not row:
                continue
            yr_vals = [int(v) for v in row if isinstance(v, (int, float)) and 2015 <= v <= 2030]
            if len(yr_vals) >= 3:
                year_row_i = ri
                seen_yrs: list = []
                seen_set: set = set()
                for v in row:
                    if isinstance(v, (int, float)) and 2015 <= v <= 2030:
                        yr = int(v)
                        if yr not in seen_set:
                            seen_yrs.append(yr)
                            seen_set.add(yr)
                years = sorted(seen_yrs)
                break

        if year_row_i >= 0:
            brand_names_row_i = year_row_i - 2
            data_start_row_i = year_row_i + 1
            n_years = len(years)

            # Map brands to starting columns via display-name matching in brand_names_row
            if 0 <= brand_names_row_i < len(pen_rows):
                bn_row = pen_rows[brand_names_row_i]
                for ci, v in enumerate(bn_row):
                    if not (isinstance(v, str) and v.strip()):
                        continue
                    v_s = v.strip()
                    for bk in brands_order:
                        if bk not in brand_col_starts and (
                            v_s == brand_display[bk] or
                            v_s in brand_display[bk] or
                            brand_display[bk] in v_s
                        ):
                            brand_col_starts[bk] = ci
                            break

            # Positional fallback: assign remaining brands by string column order
            if len(brand_col_starts) < len(brands_order) and 0 <= brand_names_row_i < len(pen_rows):
                str_cols = [ci for ci, v in enumerate(pen_rows[brand_names_row_i])
                            if isinstance(v, str) and v.strip()]
                assigned = 0
                for bk in brands_order:
                    if bk not in brand_col_starts and assigned < len(str_cols):
                        brand_col_starts[bk] = str_cols[assigned]
                    assigned += 1

            # Scan data rows into segment groups.
            # Structure: col 0 = section header (Пол/Возраст/…) on first row of each group, else None.
            #            col 1 = individual row label (Мужчины, 18-24, Большая Москва, …).
            #            Total row: col 0 = 'Total', col 1 = None.
            current_seg = 'total'

            for ri in range(data_start_row_i, len(pen_rows)):
                row = pen_rows[ri]
                if not row:
                    continue

                col0 = row[0] if len(row) > 0 else None
                col1 = row[1] if len(row) > 1 else None

                # Col 0 section header → switch current segment group
                if col0 is not None:
                    col0_lower = str(col0).strip().lower()
                    for kw, seg_key in _PEN_SEG_KEYWORDS.items():
                        if kw in col0_lower:
                            current_seg = seg_key
                            break

                # Label: prefer col 1; fall back to col 0 (Total row has col 1 = None)
                if col1 is not None and str(col1).strip():
                    label_str = str(col1).strip()
                elif col0 is not None and str(col0).strip():
                    label_str = str(col0).strip()
                else:
                    label_str = ''

                if not label_str:
                    continue

                # Total row: col 0 = 'Total', ensure correct segment
                if label_str.lower() == 'total':
                    current_seg = 'total'

                # Check presence of any non-null brand data
                has_data = False
                for bk in brands_order:
                    b_col = brand_col_starts.get(bk)
                    if b_col is None:
                        continue
                    for yi in range(n_years):
                        idx = b_col + yi
                        if idx < len(row) and row[idx] is not None:
                            has_data = True
                            break
                    if has_data:
                        break

                if not has_data:
                    continue

                e: dict = {'label': label_str}
                for bk in brands_order:
                    b_col = brand_col_starts.get(bk)
                    if b_col is not None:
                        e[bk] = [_cv(row[b_col + yi] if b_col + yi < len(row) else None)
                                  for yi in range(n_years)]
                    else:
                        e[bk] = [None] * n_years
                pen_segments[current_seg].append(e)

    generated = months_sorted[-1] if months_sorted else ''
    return {
        'format': 'monthly',
        'generated': generated,
        'brands': brands_order,
        'brandNames': brand_display,
        'metricKeys': metric_keys,
        'metricLabels': metric_labels_out,
        'metrics': metrics,
        'penetration': {
            'brands': brands_order,
            'years': years,
            'segments': pen_segments,
        },
    }


def process(file_path: str) -> dict:
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)

    if _is_monthly_format(wb):
        result = _process_monthly_auto(wb)
        wb.close()
        return result

    # Mineral water format (quarterly)
    rows0 = list(wb.worksheets[1].iter_rows(values_only=True))  # по гендеру
    rows1 = list(wb.worksheets[0].iter_rows(values_only=True))  # по возрастам
    rows2 = list(wb.worksheets[2].iter_rows(values_only=True))  # частота
    wb.close()

    metrics = {}
    for metric_name, brand_rows in METRIC_ROWS.items():
        series = []
        for q, col in CHRONO:
            entry = {'quarter': q}
            for brand in BRANDS:
                ri = brand_rows.get(brand)
                entry[brand] = vf(rows0[ri][col]) if ri is not None else 0.0
            series.append(entry)
        metrics[metric_name] = series

    funnel = {}
    for q, col in CHRONO:
        pop = vf(rows0[TOTAL_ROW_S0][col])
        entries = []
        for brand in BRANDS:
            row = {'brand': brand, 'brandName': BRAND_DISPLAY[brand], 'population': pop}
            for metric_name, brand_rows in METRIC_ROWS.items():
                ri = brand_rows.get(brand)
                raw = vf(rows0[ri][col]) if ri is not None else 0.0
                row[metric_name] = round(raw / pop * 100, 2) if pop > 0 else 0.0
            entries.append(row)
        funnel[q] = entries

    pen_by_age = {}
    for qi, q in enumerate(QUARTERS):
        entries = []
        for age in AGE_GROUPS:
            cb = AGE_COL_START[age]
            entry = {'age': age, 'population': vf(rows1[AGE_TOTAL_ROW][cb + qi])}
            for brand, ri in CONSUMPTION_ROWS.items():
                entry[brand] = vf(rows1[ri][cb + qi])
            entries.append(entry)
        pen_by_age[q] = entries
    for cur, prev in zip(pen_by_age[CURRENT_Q], pen_by_age[PRIOR_Q]):
        for brand in BRANDS:
            cur[f'{brand}Delta'] = round(cur[brand] - prev[brand], 1)

    freq_by_brand = {}
    for qi, q in enumerate(QUARTERS):
        pop = vf(rows2[FREQ_TOTAL_ROW][FREQ_COL_START['total'] + qi])
        entries = []
        for brand, ri in FREQ_ROWS.items():
            entry = {
                'brand': brand, 'brandName': BRAND_DISPLAY[brand],
                'population': pop, 'total': vf(rows2[ri][FREQ_COL_START['total'] + qi]),
            }
            for cat in FREQ_CATS:
                entry[cat] = vf(rows2[ri][FREQ_COL_START[cat] + qi])
            entries.append(entry)
        freq_by_brand[q] = entries
    prior_totals = {e['brand']: e['total'] for e in freq_by_brand[PRIOR_Q]}
    for e in freq_by_brand[CURRENT_Q]:
        e['totalDelta'] = round(e['total'] - prior_totals.get(e['brand'], e['total']), 1)

    return {
        'generated': CURRENT_Q, 'brands': BRANDS, 'brandNames': BRAND_DISPLAY,
        'quarters': QUARTERS, 'currentQuarter': CURRENT_Q,
        'metricKeys': METRIC_KEYS, 'metricLabels': METRIC_LABELS,
        'metrics': metrics, 'funnel': funnel,
        'ageGroups': AGE_GROUPS, 'freqCats': FREQ_CATS, 'freqLabels': FREQ_LABELS,
        'penetrationByAge': pen_by_age, 'frequencyByBrand': freq_by_brand,
    }
