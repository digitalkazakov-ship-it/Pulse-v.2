import openpyxl
from collections import defaultdict

RU_MONTHS_SHORT = {
    1:'Янв', 2:'Фев', 3:'Мар', 4:'Апр', 5:'Май', 6:'Июн',
    7:'Июл', 8:'Авг', 9:'Сен', 10:'Окт', 11:'Ноя', 12:'Дек',
}
DEFAULT_BRANDS = ['СЕЛО ЗЕЛЕНОЕ', 'БЕЛЕБЕЕВСКИЙ', 'БРЕСТ-ЛИТОВСК', 'СЫРОБОГАТОВ', 'ФЕТАКСА']


def _fv(v, decimals=0):
    if isinstance(v, (int, float)):
        return round(float(v), decimals) if decimals else int(round(float(v)))
    return None


def _year_ranges(row):
    """Return sorted list of (year, start_col, end_col) from a header row."""
    year_starts = []
    for ci, val in enumerate(row):
        if isinstance(val, (int, float)) and 2000 < int(val) < 2100:
            year_starts.append((int(val), ci))
    year_starts.sort()
    ranges = []
    for i, (yr, sc) in enumerate(year_starts):
        ec = year_starts[i + 1][1] - 1 if i + 1 < len(year_starts) else len(row) - 1
        ranges.append((yr, sc, ec))
    return ranges


def _parse_seasonality(ws):
    rows = list(ws.iter_rows(values_only=True))
    yr_ranges = _year_ranges(rows[1])
    years = [yr for yr, _, _ in yr_ranges]

    brands = []
    brand_data = {}
    for r in rows[3:]:
        brand = r[1]
        if not brand:
            continue
        brands.append(str(brand))
        brand_data[str(brand)] = {}
        for yr, sc, ec in yr_ranges:
            brand_data[str(brand)][str(yr)] = [_fv(r[ci]) for ci in range(sc, ec + 1)]

    data = {}
    for yr, sc, ec in yr_ranges:
        n = ec - sc + 1
        series = []
        for m in range(n):
            pt = {'month': RU_MONTHS_SHORT[m + 1]}
            for b in brands:
                vals = brand_data[b].get(str(yr), [])
                pt[b] = vals[m] if m < len(vals) else None
            series.append(pt)
        data[str(yr)] = series

    return {'brands': brands, 'defaultBrands': DEFAULT_BRANDS, 'years': years, 'data': data}


_EXCLUDED_BRANDS = {'Budgets, All media', 'Grand Total', 'Grand total'}


def _parse_regionality(ws):
    rows = list(ws.iter_rows(values_only=True))
    yr_ranges = _year_ranges(rows[1])
    years = [yr for yr, _, _ in yr_ranges]

    brands_set = set()
    media_set = set()
    regions_by_media: dict = {}        # media → ordered list of regions
    # agg[yr][media][region][brand] = total spend
    agg: dict = {str(yr): defaultdict(lambda: defaultdict(lambda: defaultdict(float)))
                 for yr, _, _ in yr_ranges}

    # Monthly totals per brand, summed across ALL media + regions ("All media" view).
    # Column ranges can be partial for the most recent year (e.g. Jan-Jul only) —
    # preserved here (unlike agg above) so the frontend can detect that and show
    # a matching same-months-prior-year comparison column.
    year_months = {str(yr): ec - sc + 1 for yr, sc, ec in yr_ranges}
    # NOTE: `n=year_months[str(yr)]` default-arg binds the value at lambda-creation
    # time — without it, the closure would capture `yr` by reference and every
    # year's defaultdict would size its lists using whatever `yr` ends up being
    # after the loop finishes (a classic late-binding bug).
    monthly_agg: dict = {str(yr): defaultdict(lambda n=year_months[str(yr)]: [0.0] * n)
                          for yr, _, _ in yr_ranges}

    # Monthly totals per brand, kept SEPARATE per media type but summed across
    # regions — for a chart that cares about (brand, media, month) and ignores
    # region entirely. monthly_by_media[media][yr_str][brand] = [v1..vN].
    # Built with plain dicts (not defaultdict-with-lambda) to sidestep the same
    # late-binding closure bug that hit monthly_agg above.
    monthly_by_media: dict = {}

    for r in rows[3:]:
        brand, media, region = r[1], r[2], r[3]
        if not brand or not media or not region:
            continue
        brand, media, region = str(brand), str(media), str(region)
        if brand in _EXCLUDED_BRANDS:
            continue
        brands_set.add(brand)
        media_set.add(media)
        if media not in regions_by_media:
            regions_by_media[media] = []
        if region not in regions_by_media[media]:
            regions_by_media[media].append(region)
        for yr, sc, ec in yr_ranges:
            total = sum(v for ci in range(sc, ec + 1) if isinstance((v := r[ci]), (int, float)))
            agg[str(yr)][media][region][brand] += total
            yr_str = str(yr)
            month_vals = monthly_agg[yr_str][brand]

            media_years = monthly_by_media.setdefault(media, {})
            media_year_brands = media_years.setdefault(yr_str, {})
            media_month_vals = media_year_brands.setdefault(brand, [0.0] * year_months[yr_str])

            for idx, ci in enumerate(range(sc, ec + 1)):
                v = r[ci]
                if isinstance(v, (int, float)):
                    month_vals[idx] += v
                    media_month_vals[idx] += v

    brands = sorted(brands_set)
    media_types = sorted(media_set)

    # Build chart data: yr → media → region → [{brand, value}]
    # Include synthetic 'Все' region (sum across all regions)
    data: dict = {}
    for yr_str, media_dict in agg.items():
        data[yr_str] = {}
        for media, region_dict in media_dict.items():
            data[yr_str][media] = {}
            totals: dict = defaultdict(float)
            for region, brand_dict in region_dict.items():
                data[yr_str][media][region] = [
                    {'brand': b, 'value': int(round(brand_dict[b]))}
                    for b in brands if brand_dict.get(b, 0) > 0
                ]
                for b, v in brand_dict.items():
                    totals[b] += v
            data[yr_str][media]['Все'] = [
                {'brand': b, 'value': int(round(totals[b]))}
                for b in brands if totals.get(b, 0) > 0
            ]

    # Prepend 'Все' to each media's region list
    regions = {media: ['Все'] + regs for media, regs in regions_by_media.items()}

    # monthlyTotal[year] = [{'month': 'Янв', brand: value, ...}, ...] — same
    # brand-totals as `data[year][media]['Все']` above, but per month, summed
    # across ALL media too, so a partial year (e.g. Jan-Jul) is visible as such
    # instead of silently rolling up into a misleadingly "full-looking" year sum.
    monthly_total: dict = {}
    for yr_str, n in year_months.items():
        series = []
        for m in range(n):
            pt: dict = {'month': RU_MONTHS_SHORT[m + 1]}
            for b in brands:
                vals = monthly_agg[yr_str].get(b)
                pt[b] = int(round(vals[m])) if vals else 0
            series.append(pt)
        monthly_total[yr_str] = series

    # monthlyByMedia[media][year] = [{'month': 'Янв', brand: value, ...}, ...] —
    # same shape as monthlyTotal, but kept separate per media type (region still
    # summed away). Lets the frontend offer a per-media-type channel switcher
    # over month-by-month brand budgets.
    monthly_by_media_out: dict = {}
    for media, yr_dict in monthly_by_media.items():
        monthly_by_media_out[media] = {}
        for yr_str, brand_dict in yr_dict.items():
            n = year_months[yr_str]
            series = []
            for m in range(n):
                pt: dict = {'month': RU_MONTHS_SHORT[m + 1]}
                for b in brands:
                    vals = brand_dict.get(b)
                    pt[b] = int(round(vals[m])) if vals else 0
                series.append(pt)
            monthly_by_media_out[media][yr_str] = series

    return {
        'brands': brands,
        'mediaTypes': media_types,
        'years': years,
        'regions': regions,
        'data': data,
        'monthlyTotal': monthly_total,
        'monthlyByMedia': monthly_by_media_out,
    }


def _parse_trp(ws4, ws5):
    def read_ws(ws):
        rows = list(ws.iter_rows(values_only=True))
        header = rows[1]
        periods = [str(int(h)) if isinstance(h, (int, float)) else str(h) for h in header[2:6] if h is not None]
        period_cols = list(range(2, 2 + len(periods)))
        brands, data = [], {}
        for r in rows[2:]:
            brand = r[1]
            if not brand or str(brand) == 'Grand Total':
                continue
            brands.append(str(brand))
            data[str(brand)] = {p: _fv(r[ci], 1) for p, ci in zip(periods, period_cols)}
        return brands, periods, data

    brands20, periods20, data20 = read_ws(ws4)
    brands_s, _, data_s = read_ws(ws5)
    all_brands = list(dict.fromkeys(brands20 + brands_s))
    return {'brands': all_brands, 'periods': periods20, 'trp20': data20, 'trps': data_s}


def _parse_tv_strategy(ws):
    rows = list(ws.iter_rows(values_only=True))
    brand_row, placement_row = rows[1], rows[2]

    # Map col index → (brand, placement)
    col_map = {}
    current_brand = None
    for ci, val in enumerate(brand_row):
        if val:
            current_brand = str(val)
        if current_brand and ci >= 3 and placement_row[ci]:
            col_map[ci] = (current_brand, str(placement_row[ci]))

    brands = list(dict.fromkeys(b for b, _ in col_map.values()))
    placements = list(dict.fromkeys(p for _, p in col_map.values()))

    # data[yr_str][placement][week] = {brand: value}
    data = defaultdict(lambda: defaultdict(lambda: defaultdict(dict)))
    current_year = None
    for r in rows[3:]:
        if r[1] is not None:
            current_year = int(r[1])
        week = r[2]
        if current_year is None or week is None:
            continue
        yr_str = str(current_year)
        w = int(week)
        for ci, (brand, placement) in col_map.items():
            val = _fv(r[ci], 2)
            if val is not None and val > 0:
                data[yr_str][placement][w][brand] = val

    years = sorted(data.keys())
    chart_data = {}
    for yr_str in years:
        chart_data[yr_str] = {}
        for placement in placements:
            week_dict = data[yr_str].get(placement, {})
            if not week_dict:
                continue
            max_week = max(week_dict.keys())
            series = [{'week': w, **{b: week_dict.get(w, {}).get(b) for b in brands}}
                      for w in range(1, max_week + 1)]
            chart_data[yr_str][placement] = series

    return {'brands': brands, 'placements': placements, 'years': [int(y) for y in years], 'data': chart_data}


def _parse_clip_duration(ws):
    rows = list(ws.iter_rows(values_only=True))
    header = rows[2]
    periods, period_cols = [], []
    for ci, val in enumerate(header):
        if ci >= 3 and val is not None and ci - 3 < 3:
            periods.append(str(int(val)) if isinstance(val, (int, float)) else str(val))
            period_cols.append(ci)

    brands, clip_data, durations_set = [], {}, set()
    current_brand = None
    for r in rows[3:]:
        if r[1]:
            current_brand = str(r[1])
            if current_brand not in clip_data:
                brands.append(current_brand)
                clip_data[current_brand] = {}
        if current_brand is None or r[2] is None:
            continue
        dur = int(r[2])
        durations_set.add(dur)
        clip_data[current_brand][dur] = {
            p: round(float(r[ci]) * 100, 1) if isinstance(r[ci], (int, float)) else None
            for p, ci in zip(periods, period_cols)
        }

    durations = sorted(durations_set)
    chart_data = {}
    for period in periods:
        chart_data[period] = [
            {'brand': b, **{str(d): clip_data[b].get(d, {}).get(period) for d in durations}}
            for b in brands
        ]

    return {'brands': brands, 'durations': durations, 'periods': periods, 'data': chart_data}


def _parse_media_mix(ws) -> dict:
    """Sheet 1: Media Mix by brand and channel for multiple periods (rubles → млн руб)."""
    rows = list(ws.iter_rows(values_only=True))
    # Row 1: [None, 'Brand', 'Media', period1, period2, ...]
    header = rows[1]
    periods = [str(v) if not isinstance(v, (int, float)) else str(int(v))
               for v in header[3:] if v is not None]
    period_cols = [ci for ci, v in enumerate(header) if v is not None and ci >= 3]

    _SKIP = {'budgets, all media', 'grand total'}
    brands: list[str] = []
    data: dict[str, dict[str, dict[str, float]]] = {}  # brand → period → channel → млн руб

    current_brand: str | None = None
    for row in rows[2:]:
        b, ch = row[1], row[2]
        if isinstance(b, str) and b.strip():
            bk = b.strip()
            if bk.lower() in _SKIP:
                current_brand = None
                continue
            current_brand = bk
            if bk not in data:
                brands.append(bk)
                data[bk] = {}
        if current_brand is None or not isinstance(ch, str) or not ch.strip():
            continue
        ch_key = ch.strip().upper()
        for i, p in enumerate(periods):
            if i >= len(period_cols):
                break
            v = row[period_cols[i]]
            if isinstance(v, (int, float)) and v:
                data[current_brand].setdefault(p, {})[ch_key] = round(float(v) / 1_000_000, 2)

    return {'brands': brands, 'periods': periods, 'data': data}


def _parse_aww_direct(ws) -> dict:
    """Sheet 7: Direct TV only (excl. sponsorship) — TRPs, active weeks, AWW per brand per period."""
    rows = list(ws.iter_rows(values_only=True))
    header = rows[1]
    brands = [str(v) for v in header[3:] if v is not None]
    brand_cols = [ci for ci, v in enumerate(header) if v is not None and ci >= 3]

    periods: list[str] = []
    data: dict = {}
    current_period: str | None = None

    for row in rows[2:]:
        col1, metric = row[1], row[2]
        # New period only when col2 == 'Total TRPs' — avoids picking up sub-labels and footer tables
        if metric == 'Total TRPs' and col1 is not None:
            current_period = str(int(col1)) if isinstance(col1, (int, float)) else str(col1).strip()
            if current_period not in data:
                periods.append(current_period)
                data[current_period] = {'trp': {}, 'weeks': {}, 'aww': {}}
        if metric is None or current_period is None:
            continue
        for j, b in enumerate(brands):
            if j >= len(brand_cols):
                break
            v = row[brand_cols[j]] if brand_cols[j] < len(row) else None
            if not isinstance(v, (int, float)):
                continue
            if metric == 'Total TRPs':
                data[current_period]['trp'][b] = round(float(v), 1)
            elif metric == 'Weeks qty':
                data[current_period]['weeks'][b] = int(v)
            elif metric == 'AWW':
                data[current_period]['aww'][b] = round(float(v), 1)

    return {'brands': brands, 'periods': periods, 'data': data}


def _parse_tv_sponsorship(ws) -> dict:
    """Sheet 10: TV sponsorship TRPs per brand per year (monthly columns)."""
    rows = list(ws.iter_rows(values_only=True))

    # Row 1: year values at their start columns
    year_starts: list[tuple[int, int]] = sorted(
        (int(v), ci) for ci, v in enumerate(rows[1])
        if isinstance(v, (int, float)) and 2020 <= v <= 2030
    )
    years = [yr for yr, _ in year_starts]

    def _year_for_col(ci: int) -> int | None:
        yr_cur = None
        for yr, start in year_starts:
            if ci >= start:
                yr_cur = yr
            else:
                break
        return yr_cur

    _SKIP_KW = {'nat tv', 'trps w', 'тип ролика', 'grand total', 'budgets'}
    brands_with_sponsorship: list[str] = []
    brand_trp: dict[str, dict[int, float]] = {}  # brand → {year → total TRP}

    for row in rows[3:]:
        b = row[1]
        if not (isinstance(b, str) and b.strip()):
            continue
        b = b.strip()
        if b.endswith(' Total'):
            pass  # handled below
        elif not any(kw in b.lower() for kw in _SKIP_KW):
            if b not in brands_with_sponsorship:
                brands_with_sponsorship.append(b)
            continue
        else:
            continue
        if not b.endswith(' Total'):
            continue
        brand_name = b[: -len(' Total')]
        yearly: dict[int, float] = {}
        for ci, v in enumerate(row):
            if not isinstance(v, (int, float)) or ci < 5:
                continue
            yr = _year_for_col(ci)
            if yr:
                yearly[yr] = yearly.get(yr, 0.0) + v
        brand_trp[brand_name] = {yr: round(tot, 1) for yr, tot in yearly.items()}

    return {'brands': brands_with_sponsorship, 'years': years, 'trp': brand_trp}


def process(file_path: str) -> dict:
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    result = {
        'mediaMix':      _parse_media_mix(wb.worksheets[1]),
        'seasonality':   _parse_seasonality(wb.worksheets[2]),
        'regionality':   _parse_regionality(wb.worksheets[3]),
        'trp':           _parse_trp(wb.worksheets[4], wb.worksheets[5]),
        'tvStrategy':    _parse_tv_strategy(wb.worksheets[6]),
        'awwDirect':     _parse_aww_direct(wb.worksheets[7]),
        'clipDuration':  _parse_clip_duration(wb.worksheets[8]),
        'tvSponsorship': _parse_tv_sponsorship(wb.worksheets[10]),
    }
    wb.close()
    return result
