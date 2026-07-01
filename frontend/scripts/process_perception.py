"""
Converts SL.xlsx to perception.json for the Perception & Social page.

Usage:
    py scripts/process_perception.py

Input:
    FILE — Social listening/SL.xlsx
    Sheet "Шаг 4.1 - SL"   → 3 SL category charts + sentiment
    Sheet "Шаг 4.4 - Имидж" → image characteristics + brand×image matrix

Sentiment history accumulates across runs: each run upserts current GENERATED month.

Output:
    public/data/perception.json
"""

import openpyxl
import json
from pathlib import Path

FILE   = r'C:\Users\ebone\.claude\New folder\Pulse\Files\Social listening\SL.xlsx'
OUTPUT = Path(__file__).parent.parent / 'public' / 'data' / 'perception.json'

BRANDS = ['Narzan', 'Borjomi', 'Senezhskaya', 'Svyatoy']
BRAND_DISPLAY = {
    'Narzan':      'Нарзан',
    'Borjomi':     'Боржоми',
    'Senezhskaya': 'Сенежская',
    'Svyatoy':     'Святой источник',
}

# Month when data was collected — update each period
GENERATED = '2026-05'

RU_MONTHS = {
    1:'Янв', 2:'Фев', 3:'Мар', 4:'Апр', 5:'Май', 6:'Июн',
    7:'Июл', 8:'Авг', 9:'Сен', 10:'Окт', 11:'Ноя', 12:'Дек',
}

def month_label(ym: str) -> str:
    y, m = int(ym[:4]), int(ym[5:7])
    label = RU_MONTHS[m]
    if y != 2026:
        label += f" '{str(y)[2:]}"
    return label


# ── Sheet "Шаг 4.1 - SL" ─────────────────────────────────────────────────────
#
# ПОЗИТИВ абс (cols D–I, 0-based 3–8):
#   Нарзан=3, Рычал-су=4(skip), Свят.ист=5, Сенежская=6, Аква М.=7(skip), Боржоми=8
# НЕГАТИВ абс (cols P–U, 0-based 15–20):
#   Нарзан=15, Рычал-су=16(skip), Свят.ист=17, Сенежская=18, Аква М.=19(skip), Боржоми=20

POS_COLS = {'Narzan': 3, 'Svyatoy': 5, 'Senezhskaya': 6, 'Borjomi': 8}
NEG_COLS = {'Narzan': 15, 'Svyatoy': 17, 'Senezhskaya': 18, 'Borjomi': 20}

SL_CATEGORIES = [
    {
        'key': 'assortment',
        'label': 'Ассортимент',
        'params': [
            {'key': 'p0', 'label': 'Уникальные и эксклюзивные продукты', 'row': 4},
            {'key': 'p1', 'label': 'Широкий ассортимент', 'row': 5},
            {'key': 'p2', 'label': 'Высокое качество товаров', 'row': 6},
            {'key': 'p3', 'label': 'Высокое качество упаковки', 'row': 7},
        ],
    },
    {
        'key': 'price',
        'label': 'Цена',
        'params': [
            {'key': 'p0', 'label': 'Конкурентоспособные и выгодные цены', 'row': 8},
            {'key': 'p1', 'label': 'Наличие скидок и промо', 'row': 9},
        ],
    },
    {
        'key': 'product',
        'label': 'Продукт',
        'params': [
            {'key': 'p0', 'label': 'Приятный вкус', 'row': 13},
            {'key': 'p1', 'label': 'Польза от употребления', 'row': 14},
            {'key': 'p2', 'label': 'Можно часто пить', 'row': 15},
        ],
    },
]


def int_val(v) -> int:
    return int(v) if isinstance(v, (int, float)) else 0


def read_sl(rows: list) -> tuple[dict, dict]:
    sl = {}
    pos_total = {b: 0 for b in BRANDS}
    neg_total = {b: 0 for b in BRANDS}

    for cat in SL_CATEGORIES:
        chart_data: dict[str, list] = {}
        cat_pos = {b: 0 for b in BRANDS}
        cat_neg = {b: 0 for b in BRANDS}

        for p in cat['params']:
            row = rows[p['row']]
            pts = []
            for b in BRANDS:
                pv = int_val(row[POS_COLS[b]])
                nv = int_val(row[NEG_COLS[b]])
                pts.append({'brand': b, 'brandName': BRAND_DISPLAY[b], 'pos': pv, 'neg': nv})
                cat_pos[b] += pv
                cat_neg[b] += nv
            chart_data[p['key']] = pts

        chart_data['total'] = [
            {'brand': b, 'brandName': BRAND_DISPLAY[b], 'pos': cat_pos[b], 'neg': cat_neg[b]}
            for b in BRANDS
        ]
        for b in BRANDS:
            pos_total[b] += cat_pos[b]
            neg_total[b] += cat_neg[b]

        sl[cat['key']] = {
            'label': cat['label'],
            'params': [{'key': 'total', 'label': 'Все параметры'}] + [
                {'key': p['key'], 'label': p['label']} for p in cat['params']
            ],
            'chartData': chart_data,
        }

    sentiment = {
        b: round((pos_total[b] - neg_total[b]) / (pos_total[b] + neg_total[b]) * 100, 1)
        if (pos_total[b] + neg_total[b]) > 0 else 0.0
        for b in BRANDS
    }
    return sl, sentiment


# ── Sheet "Шаг 4.4 - Имидж" ──────────────────────────────────────────────────
#
# Table 1 — absolute (brand cols 0-based):
#   Нарзан=2, Боржоми=3, Свят.ист=4, Рычал-су=5(skip), Сенежская=6, Аква М.=7(skip)
# Table 3 — vs median (brand cols 0-based):
#   Нарзан=22, Рычал-су=23(skip), Свят.ист=24, Сенежская=25, Аква М.=26(skip), Боржоми=27
#
# Characteristic labels in col B (index 1), rows 2–21 (20 total)

ABS_COLS = {'Narzan': 2, 'Borjomi': 3, 'Svyatoy': 4, 'Senezhskaya': 6}
MED_COLS = {'Narzan': 22, 'Svyatoy': 24, 'Senezhskaya': 25, 'Borjomi': 27}

ALL_ROWS = list(range(2, 22))  # 20 characteristics
# 7 selected for Имиджевые характеристики chart:
# Надежный(2), Доступный(4), Дорогой(5), Современный(8), Натуральный(10), Вкусный(18), Лечебный(20)
CHAR_ROWS = [2, 4, 5, 8, 10, 18, 20]


def read_image(rows: list) -> dict:
    def abs_point(r: int) -> dict:
        row = rows[r]
        pt = {'label': str(row[1])}
        for b in BRANDS:
            pt[b] = int_val(row[ABS_COLS[b]])
        return pt

    def med_point(r: int) -> dict:
        row = rows[r]
        pt = {'label': str(row[1])}
        for b in BRANDS:
            v = rows[r][MED_COLS[b]]
            pt[b] = round(float(v) * 100, 1) if isinstance(v, (int, float)) else 0.0
        return pt

    return {
        'imageChars': [abs_point(r) for r in CHAR_ROWS],
        'matrix': {
            'absolute': [abs_point(r) for r in ALL_ROWS],
            'median':   [med_point(r) for r in ALL_ROWS],
        },
    }


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    print("Reading Excel file...")
    wb = openpyxl.load_workbook(FILE, read_only=True, data_only=True)
    ws_sl    = wb['Шаг 4.1 - SL']
    ws_image = wb['Шаг 4.4 - Имидж']
    rows_sl    = list(ws_sl.iter_rows(values_only=True))
    rows_image = list(ws_image.iter_rows(values_only=True))
    wb.close()

    sl_data, sentiment = read_sl(rows_sl)
    image_data = read_image(rows_image)

    # Accumulate sentiment history: read existing, upsert current month
    history = []
    if OUTPUT.exists():
        with open(OUTPUT, encoding='utf-8') as f:
            existing = json.load(f)
        history = existing.get('sentiment', [])

    lbl = month_label(GENERATED)
    entry = {'period': GENERATED, 'month': lbl, **sentiment}
    history = [h for h in history if h.get('period') != GENERATED]
    history.append(entry)
    history.sort(key=lambda h: h.get('period', ''))

    output = {
        'generated': GENERATED,
        'brands': BRANDS,
        'brandNames': BRAND_DISPLAY,
        'sl': sl_data,
        'sentiment': history,
        **image_data,
    }

    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    with open(OUTPUT, 'w', encoding='utf-8') as f:
        json.dump(output, f, ensure_ascii=False, indent=2)

    print(f"\nWritten: {OUTPUT}")

    print(f"\nSentiment ({lbl}):")
    for b in BRANDS:
        print(f"  {BRAND_DISPLAY[b]:20s}  {sentiment[b]:+.1f}%")

    print("\nSL Totals (Все параметры):")
    for cat in SL_CATEGORIES:
        print(f"\n  {cat['label']}:")
        for pt in sl_data[cat['key']]['chartData']['total']:
            print(f"    {pt['brandName']:20s}  pos={pt['pos']:5d}  neg={pt['neg']:5d}")

    print(f"\nImage: {len(image_data['imageChars'])} chars, matrix {len(image_data['matrix']['absolute'])} rows")


if __name__ == '__main__':
    main()
