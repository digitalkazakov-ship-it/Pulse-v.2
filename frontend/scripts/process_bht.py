"""
Converts Минеральные воды _ Джава.xlsx to bht.json for BrandAwareness page.

Input sheets (0-based index):
  0  "по гендеру"    — Brand metrics (Total only, columns 2-5)
       Row 10: Total population  (col 2=Q1 2026, 3=Q4 2025, 4=Q3 2025, 5=Q2 2025)
       Metric blocks (skip Aqua Minerale and Нарзан-Су rows):
         Первое упоминание:    Borjomi=12, Narzan=13, Svyatoy=15, Senezhskaya=16
         Спонтанное знание:    Borjomi=19, Narzan=20, Svyatoy=22, Senezhskaya=23
         Подсказанное знание:  Borjomi=26, Narzan=27, Svyatoy=29, Senezhskaya=30
         Потребление:          Borjomi=33, Narzan=34, Svyatoy=36, Senezhskaya=37
         Рассмотрение:         Borjomi=47, Narzan=48, Svyatoy=50, Senezhskaya=51

  1  "по возрастам"  — Penetration by age group
       Row 9: Total population by age group
       Rows 31-36: "Потребление" block
         31=Aqua Minerale (skip), 32=Borjomi, 33=Narzan,
         34=Нарзан-Су (skip), 35=Svyatoy, 36=Senezhskaya
       Age group column starts: Total=2, 12-17=6, 18-24=10, 25-34=14,
                                 35-44=18, 45-54=22, 55-64=26
       Within each group: offset 0=Q1 2026, 1=Q4 2025, 2=Q3 2025, 3=Q2 2025

  2  "частота"       — Frequency of consumption
       Row 9: Total population (same offsets as above)
       Rows 10-15: 10=Aqua Minerale (skip), 11=Borjomi, 12=Narzan,
                   13=Нарзан-Су (skip), 14=Svyatoy, 15=Senezhskaya
       Category column starts: total=2, daily=6, week23=10,
                                week1=14, month23=18, less=22

Output:
  public/data/bht.json
"""

import openpyxl
import json
from pathlib import Path

FILE   = r'C:\Users\ebone\.claude\New folder\Pulse\Files\BHT\Минеральные воды _ Джава.xlsx'
OUTPUT = Path(__file__).parent.parent / 'public' / 'data' / 'bht.json'

BRANDS = ['Narzan', 'Borjomi', 'Senezhskaya', 'Svyatoy']
BRAND_DISPLAY = {
    'Narzan':      'Нарзан',
    'Borjomi':     'Боржоми',
    'Senezhskaya': 'Сенежская',
    'Svyatoy':     'Святой источник',
}

QUARTERS   = ['Q1 2026', 'Q4 2025', 'Q3 2025', 'Q2 2025']
CURRENT_Q  = 'Q1 2026'
PRIOR_Q    = 'Q4 2025'

# ── Sheet 0: brand metrics ────────────────────────────────────────────────────
TOTAL_ROW_S0 = 10
METRIC_ROWS = {
    'topOfMind':     {'Borjomi': 12, 'Narzan': 13, 'Svyatoy': 15, 'Senezhskaya': 16},
    'spontaneous':   {'Borjomi': 19, 'Narzan': 20, 'Svyatoy': 22, 'Senezhskaya': 23},
    'aided':         {'Borjomi': 26, 'Narzan': 27, 'Svyatoy': 29, 'Senezhskaya': 30},
    'consumption':   {'Borjomi': 33, 'Narzan': 34, 'Svyatoy': 36, 'Senezhskaya': 37},
    'consideration': {'Borjomi': 47, 'Narzan': 48, 'Svyatoy': 50, 'Senezhskaya': 51},
}
METRIC_KEYS = ['topOfMind', 'spontaneous', 'aided', 'consideration', 'consumption']
METRIC_LABELS = {
    'topOfMind':     'Первое упоминание',
    'spontaneous':   'Спонтанное',
    'aided':         'Подсказанное',
    'consideration': 'Рассмотрение',
    'consumption':   'Потребление',
}
# Column offsets for Total: col 2=Q1 2026, 3=Q4 2025, 4=Q3 2025, 5=Q2 2025
# Chronological order for chart X-axis:
CHRONO = [('Q2 2025', 5), ('Q3 2025', 4), ('Q4 2025', 3), ('Q1 2026', 2)]

# ── Sheet 1: penetration by age ───────────────────────────────────────────────
AGE_TOTAL_ROW = 9
CONSUMPTION_ROWS = {'Borjomi': 32, 'Narzan': 33, 'Svyatoy': 35, 'Senezhskaya': 36}

AGE_GROUPS = ['Total', '12-17', '18-24', '25-34', '35-44', '45-54', '55-64']
AGE_COL_START = {
    'Total': 2, '12-17': 6, '18-24': 10, '25-34': 14,
    '35-44': 18, '45-54': 22, '55-64': 26,
}

# ── Sheet 2: frequency ────────────────────────────────────────────────────────
FREQ_TOTAL_ROW = 9
FREQ_ROWS = {'Borjomi': 11, 'Narzan': 12, 'Svyatoy': 14, 'Senezhskaya': 15}

FREQ_CATS = ['daily', 'week23', 'week1', 'month23', 'less']
FREQ_LABELS = {
    'daily':   'Каждый день',
    'week23':  '2-3 раза в нед.',
    'week1':   '1 раз в нед.',
    'month23': '2-3 раза в мес.',
    'less':    'Еще реже',
}
FREQ_COL_START = {
    'total': 2, 'daily': 6, 'week23': 10, 'week1': 14, 'month23': 18, 'less': 22,
}


def vf(v):
    return round(float(v), 1) if isinstance(v, (int, float)) else 0.0


def main():
    print('Reading Excel file...')
    wb = openpyxl.load_workbook(FILE, read_only=True, data_only=True)

    # ── Sheet 0: brand metrics ────────────────────────────────────────────────
    ws0 = wb.worksheets[0]
    rows0 = list(ws0.iter_rows(values_only=True))
    print(f"  Sheet 0: '{ws0.title}', {len(rows0)} rows")

    # Metric series (chronological Q2 2025 → Q1 2026)
    metrics = {}
    for metric_name, brand_rows in METRIC_ROWS.items():
        series = []
        for q, col in CHRONO:
            entry = {'quarter': q}
            for brand in BRANDS:
                ri = brand_rows.get(brand)
                entry[brand] = vf(rows0[ri][col]) if ri is not None else 0.0
            series.append(entry)
        metrics[metric_name] = series

    # Funnel: (metric / population) × 100, per quarter
    funnel = {}
    for q, col in CHRONO:
        pop = vf(rows0[TOTAL_ROW_S0][col])
        entries = []
        for brand in BRANDS:
            row = {'brand': brand, 'brandName': BRAND_DISPLAY[brand], 'population': pop}
            for metric_name, brand_rows in METRIC_ROWS.items():
                ri = brand_rows.get(brand)
                raw = vf(rows0[ri][col]) if ri is not None else 0.0
                row[metric_name] = round(raw / pop * 100, 2) if pop > 0 else 0.0
            entries.append(row)
        funnel[q] = entries

    # ── Sheet 1: penetration by age ───────────────────────────────────────────
    ws1 = wb.worksheets[1]
    rows1 = list(ws1.iter_rows(values_only=True))
    print(f"  Sheet 1: '{ws1.title}', {len(rows1)} rows")

    pen_by_age = {}
    for qi, q in enumerate(QUARTERS):
        entries = []
        for age in AGE_GROUPS:
            cb = AGE_COL_START[age]
            entry = {'age': age, 'population': vf(rows1[AGE_TOTAL_ROW][cb + qi])}
            for brand, ri in CONSUMPTION_ROWS.items():
                entry[brand] = vf(rows1[ri][cb + qi])
            entries.append(entry)
        pen_by_age[q] = entries

    for cur, prev in zip(pen_by_age[CURRENT_Q], pen_by_age[PRIOR_Q]):
        for brand in BRANDS:
            cur[f'{brand}Delta'] = round(cur[brand] - prev[brand], 1)

    # ── Sheet 2: frequency ────────────────────────────────────────────────────
    ws2 = wb.worksheets[2]
    rows2 = list(ws2.iter_rows(values_only=True))
    print(f"  Sheet 2: '{ws2.title}', {len(rows2)} rows")

    freq_by_brand = {}
    for qi, q in enumerate(QUARTERS):
        pop = vf(rows2[FREQ_TOTAL_ROW][FREQ_COL_START['total'] + qi])
        entries = []
        for brand, ri in FREQ_ROWS.items():
            entry = {
                'brand':      brand,
                'brandName':  BRAND_DISPLAY[brand],
                'population': pop,
                'total':      vf(rows2[ri][FREQ_COL_START['total'] + qi]),
            }
            for cat in FREQ_CATS:
                entry[cat] = vf(rows2[ri][FREQ_COL_START[cat] + qi])
            entries.append(entry)
        freq_by_brand[q] = entries

    prior_totals = {e['brand']: e['total'] for e in freq_by_brand[PRIOR_Q]}
    for e in freq_by_brand[CURRENT_Q]:
        e['totalDelta'] = round(e['total'] - prior_totals.get(e['brand'], e['total']), 1)

    wb.close()

    output = {
        'generated':        '2026-05',
        'brands':           BRANDS,
        'brandNames':       BRAND_DISPLAY,
        'quarters':         QUARTERS,
        'currentQuarter':   CURRENT_Q,
        'metricKeys':       METRIC_KEYS,
        'metricLabels':     METRIC_LABELS,
        'metrics':          metrics,
        'funnel':           funnel,
        'ageGroups':        AGE_GROUPS,
        'freqCats':         FREQ_CATS,
        'freqLabels':       FREQ_LABELS,
        'penetrationByAge': pen_by_age,
        'frequencyByBrand': freq_by_brand,
    }

    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    with open(OUTPUT, 'w', encoding='utf-8') as f:
        json.dump(output, f, ensure_ascii=False, indent=2)
    print(f'\nWritten: {OUTPUT}')

    # Quick verification
    print('\nMetrics Q1 2026 (тыс. чел):')
    for mk in METRIC_KEYS:
        last = metrics[mk][-1]
        vals = '  '.join(f"{BRAND_DISPLAY[b]}={last[b]:.0f}" for b in BRANDS)
        print(f"  {METRIC_LABELS[mk]:25s}  {vals}")

    print('\nFunnel Q1 2026 (%):')
    for e in funnel[CURRENT_Q]:
        vals = '  '.join(f"{METRIC_LABELS[mk]}={e[mk]:.1f}" for mk in METRIC_KEYS)
        print(f"  {BRAND_DISPLAY[e['brand']]:22s}  {vals}")


if __name__ == '__main__':
    main()
