"""
Converts Site Stats Excel reports to digital_audit.json for the dashboard.

Usage:
    py scripts/process_digital.py

Input files (edit paths below if needed):
    FILE_CURR — current year data  (e.g. report act.xlsx,   Jan–Apr 2026)
    FILE_PREV — previous year data (e.g. report act 2.xlsx, Jan–Dec 2025)

Output:
    public/data/digital_audit.json
"""

import openpyxl
import json
from collections import defaultdict
from pathlib import Path

# ── Paths ────────────────────────────────────────────────────────────────────

FILE_CURR = r'C:\Users\ebone\.claude\New folder\Pulse\Files\report act.xlsx'
FILE_PREV = r'C:\Users\ebone\.claude\New folder\Pulse\Files\report act 2.xlsx'

OUTPUT = Path(__file__).parent.parent / 'public' / 'data' / 'digital_audit.json'

# ── Brand mapping ─────────────────────────────────────────────────────────────

BRAND_MAP = {
    'theact.ru':    'BrandX',
    'librederm.ru': 'CompA',
    'mixit.ru':     'CompB',
    'vois.ru':      'CompC',
}

BRAND_DISPLAY = {
    'BrandX': 'The Act',
    'CompA':  'Librederm',
    'CompB':  'Mixit',
    'CompC':  'Vois',
}

BRANDS = ['BrandX', 'CompA', 'CompB', 'CompC']
MAIN_BRAND = 'BrandX'

RU_MONTHS = {
    '01': 'Янв', '02': 'Фев', '03': 'Мар', '04': 'Апр',
    '05': 'Май', '06': 'Июн', '07': 'Июл', '08': 'Авг',
    '09': 'Сен', '10': 'Окт', '11': 'Ноя', '12': 'Дек',
}

# ── Helpers ───────────────────────────────────────────────────────────────────

def month_label(ym: str, show_year_if_prev=True) -> str:
    year, month = ym.split('-')
    label = RU_MONTHS[month]
    if show_year_if_prev and year != '2026':
        label += f" '{year[2:]}"
    return label


def read_daily_sheet(wb, sheet_name: str, agg: str = 'sum') -> dict:
    """Aggregate a daily sheet to monthly totals or averages per brand."""
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    header = rows[0]

    result: dict[str, dict[str, list]] = {}

    for row in rows[1:]:
        if not row[0]:
            continue
        site = str(row[0]).strip()
        if site not in BRAND_MAP:
            continue
        brand = BRAND_MAP[site]
        buckets: dict[str, list] = defaultdict(list)

        for col_i, val in enumerate(row[1:], start=1):
            if col_i >= len(header):
                break
            date_val = header[col_i]
            if date_val is None:
                continue
            ym = str(date_val)[:7]          # 'YYYY-MM'
            if isinstance(val, (int, float)) and val > 0:
                buckets[ym].append(val)

        result[brand] = {}
        for ym, vals in buckets.items():
            if agg == 'sum':
                result[brand][ym] = sum(vals)
            else:  # mean
                result[brand][ym] = sum(vals) / len(vals)

    return result


def read_unique_visits(wb) -> dict:
    """Monthly unique visitors (desktop + mobile combined) per brand."""
    ws = wb['Unique_Visits']
    rows = list(ws.iter_rows(values_only=True))
    header = rows[0]
    months = [str(h)[:7] for h in header[2:] if h is not None]

    totals: dict[str, dict[str, float]] = defaultdict(lambda: defaultdict(float))

    for row in rows[1:]:
        if not row[0]:
            continue
        site = str(row[0]).strip()
        if site not in BRAND_MAP:
            continue
        brand = BRAND_MAP[site]
        for i, ym in enumerate(months):
            val = row[2 + i]
            if isinstance(val, (int, float)) and val > 0:
                totals[brand][ym] += val

    return {b: dict(m) for b, m in totals.items()}


def read_traffic_sources(wb, sheet_name: str, device: str) -> dict:
    """
    Return traffic sources for all brands for the latest available month.
    device = 'desktop' | 'mobile'
    Returns dict[brand_key -> list of {source, organic, paid, total}]
    """
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    header = rows[0]

    # Column layout differs between sheets
    if device == 'desktop':
        # site | source_type | visit_type | date...
        date_start = 3
    else:
        # site | source_type | date...
        date_start = 2

    date_cols = [h for h in header[date_start:] if h is not None]
    if not date_cols:
        return {}
    latest_idx = date_start + len(date_cols) - 1  # 0-based column in row

    PAID_SOURCES = {'Display Ads', 'Paid Search'}

    # brand_key -> source_label -> {organic, paid}
    brand_buckets: dict[str, dict[str, dict[str, float]]] = defaultdict(
        lambda: defaultdict(lambda: {'organic': 0.0, 'paid': 0.0})
    )

    SOURCE_LABEL = {
        'Direct':         'Direct',
        'Organic Search': 'Search',
        'Paid Search':    'Search',
        'Display Ads':    'Display Ads',
        'Social Media':   'Social',
        'Referral':       'Referral',
        'Mail':           'Mail',
    }

    for row in rows[1:]:
        if not row[0]:
            continue
        site = str(row[0]).strip()
        if site not in BRAND_MAP:
            continue
        brand = BRAND_MAP[site]

        source_raw = str(row[1]).strip() if row[1] else ''
        label = SOURCE_LABEL.get(source_raw, source_raw)

        val = row[latest_idx] if latest_idx < len(row) else None
        if not isinstance(val, (int, float)) or val <= 0:
            continue

        if device == 'desktop':
            visit_type = str(row[2]).strip() if row[2] else ''
            if 'paid' in visit_type.lower() or source_raw in PAID_SOURCES:
                brand_buckets[brand][label]['paid'] += val
            else:
                brand_buckets[brand][label]['organic'] += val
        else:
            if source_raw in PAID_SOURCES:
                brand_buckets[brand][label]['paid'] += val
            else:
                brand_buckets[brand][label]['organic'] += val

    result: dict[str, list] = {}
    for brand, buckets in brand_buckets.items():
        brand_result = []
        for source, vals in buckets.items():
            total = vals['organic'] + vals['paid']
            if total > 0:
                brand_result.append({
                    'source':   source,
                    'organic':  round(vals['organic']),
                    'paid':     round(vals['paid']),
                    'total':    round(total),
                })
        brand_result.sort(key=lambda x: -x['total'])
        result[brand] = brand_result

    return result


def merge(d1: dict, d2: dict) -> dict:
    """Deep-merge two brand→month→value dicts (d1 takes priority on conflict)."""
    out: dict = {}
    for brand in set(d1) | set(d2):
        out[brand] = {}
        if brand in d2:
            out[brand].update(d2[brand])
        if brand in d1:
            out[brand].update(d1[brand])   # current-year wins
    return out


def kpi(data: dict, brand=MAIN_BRAND, curr='2026-04', prev='2026-03',
        yoy='2025-04', scale=1.0):
    bd = data.get(brand, {})
    c = bd.get(curr, 0) * scale
    p = bd.get(prev, 0) * scale
    y = bd.get(yoy,  0) * scale
    mom = round((c - p) / p * 100, 1) if p else 0
    yoy_ = round((c - y) / y * 100, 1) if y else 0
    return c, mom, yoy_


def trend(data: dict, months: list, scale=1.0, decimals=1) -> list:
    return [
        {'month': month_label(ym), **{
            b: round(data.get(b, {}).get(ym, 0) * scale, decimals)
            for b in BRANDS
        }}
        for ym in months
    ]


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    print("Reading Excel files...")
    wbc = openpyxl.load_workbook(FILE_CURR, read_only=True, data_only=True)
    wbp = openpyxl.load_workbook(FILE_PREV, read_only=True, data_only=True)

    visits  = merge(read_daily_sheet(wbc, 'Visits', 'sum'),
                    read_daily_sheet(wbp, 'Visits', 'sum'))
    unique  = merge(read_unique_visits(wbc), read_unique_visits(wbp))
    pages   = merge(read_daily_sheet(wbc, 'Pages_Per_Visit', 'mean'),
                    read_daily_sheet(wbp, 'Pages_Per_Visit', 'mean'))
    dur     = merge(read_daily_sheet(wbc, 'Avg_Visit_Duration', 'mean'),
                    read_daily_sheet(wbp, 'Avg_Visit_Duration', 'mean'))
    bounce  = merge(read_daily_sheet(wbc, 'Bounce_Rate', 'mean'),
                    read_daily_sheet(wbp, 'Bounce_Rate', 'mean'))

    td = read_traffic_sources(wbc, 'Desktop_Overview_Share', 'desktop')  # dict[brand -> list]
    tm = read_traffic_sources(wbc, 'Mobile_Overview_Share',  'mobile')   # dict[brand -> list]

    wbc.close()
    wbp.close()

    CURR, PREV, YOY = '2026-04', '2026-03', '2025-04'
    T6  = ['2025-11','2025-12','2026-01','2026-02','2026-03','2026-04']
    T12 = ['2025-05','2025-06','2025-07','2025-08','2025-09','2025-10',
           '2025-11','2025-12','2026-01','2026-02','2026-03','2026-04']

    def brand_vals(data, ym, scale=1.0, decimals=1):
        return {b: round(data.get(b, {}).get(ym, 0) * scale, decimals) for b in BRANDS}

    _, v_mom, v_yoy = kpi(visits,  scale=1/1000)
    _, u_mom, u_yoy = kpi(unique,  scale=1/1000)
    _, p_mom, p_yoy = kpi(pages)
    _, d_mom, d_yoy = kpi(dur,     scale=1/60)
    _, b_mom, b_yoy = kpi(bounce,  scale=100)

    metrics = [
        {
            'key': 'visits',  'metric': 'Monthly Visits (K)',
            **brand_vals(visits, CURR, scale=1/1000),
            'mom': v_mom, 'yoy': v_yoy,
            'trend': trend(visits, T6, scale=1/1000),
        },
        {
            'key': 'uniques', 'metric': 'Unique Visitors (K)',
            **brand_vals(unique, CURR, scale=1/1000),
            'mom': u_mom, 'yoy': u_yoy,
            'trend': trend(unique, T6, scale=1/1000),
        },
        {
            'key': 'pages',   'metric': 'Pages/Visit',
            **brand_vals(pages, CURR, decimals=2),
            'mom': p_mom, 'yoy': p_yoy,
            'trend': trend(pages, T6, decimals=2),
        },
        {
            'key': 'duration','metric': 'Avg Duration (min)',
            **brand_vals(dur, CURR, scale=1/60),
            'mom': d_mom, 'yoy': d_yoy,
            'trend': trend(dur, T6, scale=1/60),
        },
        {
            'key': 'bounce',  'metric': 'Bounce Rate (%)',
            **brand_vals(bounce, CURR, scale=100),
            'mom': b_mom, 'yoy': b_yoy,
            'trend': trend(bounce, T6, scale=100),
        },
    ]

    output = {
        'generated':             CURR,
        'brands':                BRANDS,
        'brandNames':            BRAND_DISPLAY,
        'metrics':               metrics,
        'visitsTrend':           trend(visits, T12, scale=1/1000),
        'trafficSourcesDesktop': td,
        'trafficSourcesMobile':  tm,
    }

    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    with open(OUTPUT, 'w', encoding='utf-8') as f:
        json.dump(output, f, ensure_ascii=False, indent=2)

    print(f"\nWritten: {OUTPUT}")
    print(f"\nBrandX (The Act) — {CURR}:")
    for m in metrics:
        print(f"  {m['metric']:25s} {m['BrandX']:8}  MoM {m['mom']:+.1f}%  YoY {m['yoy']:+.1f}%")


if __name__ == '__main__':
    main()
